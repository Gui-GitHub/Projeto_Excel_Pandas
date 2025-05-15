import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

# Caminho dos arquivos
caminho_pasta = r'X:\arrumar_excel'
arquivo_entrada = os.path.join(caminho_pasta, 'sc_xls_20250513150234_686_grid_db_angariacao_excel.xlsx')  # Arquivo de entrada (original)
arquivo_saida = os.path.join(caminho_pasta, 'funcionarios_dependentes_formatado.xlsx')  # Arquivo de saída formatado

# Carrega o Excel original em um DataFrame usando a biblioteca pandas.
# O pandas facilita a leitura, manipulação e análise de dados estruturados (como tabelas do Excel).
df = pd.read_excel(arquivo_entrada)

# Cria uma lista que irá armazenar os dados já organizados em um novo formato
dados_formatados = []

# Loop por todas as linhas do DataFrame. Cada linha representa um funcionário com até 6 dependentes.
# Aqui, o pandas permite acessar os dados linha a linha de forma simplificada usando .iterrows()
for _, row in df.iterrows():
    # Adiciona primeiro os dados do funcionário principal
    dados_formatados.append({
        'Nome': row['Nome Funcionario'],
        'Cpf Funcionario (repetido)': row['Cpf Funcionario'],
        'Plano Escolhido': row['Id Plano Escolhido'],
        'Sexo': '',
        'Cpf Dep': '',
        'Data Nascimento': '',
        'Nome Mae': '',
        'Parentesco': '',
        'Data Cadastro Update': row['Data Cadastro Update']
    })
    
    # Em seguida, verifica se há dados de até 6 dependentes.
    # Se houver, cria uma linha para cada dependente associando os dados relevantes.
    for i in range(1, 7):
        if pd.notna(row.get(f'Nome Dep {i}')):  # Verifica se o nome do dependente está preenchido
            data_nasc = row.get(f'Data Nascimento Dep {i}', '')
            if pd.notna(data_nasc):
                # Converte a data de nascimento para o formato brasileiro dd/mm/aaaa
                data_nasc = pd.to_datetime(data_nasc).strftime('%d/%m/%Y')
            else:
                data_nasc = ''
            dados_formatados.append({
                'Nome': row[f'Nome Dep {i}'],
                'Cpf Funcionario (repetido)': row['Cpf Funcionario'],
                'Plano Escolhido': row['Id Plano Escolhido'],
                'Sexo': row.get(f'Sexo Dep {i}', ''),
                'Cpf Dep': row.get(f'Cpf Dep {i}', ''),
                'Data Nascimento': data_nasc,
                'Nome Mae': row.get(f'Nome Mae Dep {i}', ''),
                'Parentesco': row.get(f'Parentesco Dep {i}', ''),
                'Data Cadastro Update': row['Data Cadastro Update']
            })

# Converte a lista de dicionários em um novo DataFrame formatado
df_final = pd.DataFrame(dados_formatados)

# Salva esse novo DataFrame em um arquivo Excel
df_final.to_excel(arquivo_saida, index=False)

# A partir daqui, usamos openpyxl para aplicar estilos no Excel gerado

# Carrega o arquivo Excel salvo para aplicar formatações
wb = load_workbook(arquivo_saida)
ws = wb.active

# Define estilos para o cabeçalho: cor de fundo azul e texto branco em negrito
cabecalho_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
cabecalho_font = Font(color="FFFFFF", bold=True)
alinhamento_centro = Alignment(horizontal="center", vertical="center")

# Aplica estilos às células do cabeçalho
for col in range(1, ws.max_column + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = cabecalho_fill
    celula.font = cabecalho_font
    celula.alignment = alinhamento_centro

# Aplica alinhamento centralizado a todas as demais células da planilha
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alinhamento_centro

# Salva o arquivo final com os estilos aplicados
wb.save(arquivo_saida)

print(f"Arquivo gerado com sucesso em:\n{arquivo_saida}")